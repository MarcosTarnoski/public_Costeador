# Standard library imports
import operator
from functools import reduce
from decimal import Decimal
from math import ceil
# Third party imports
from django.contrib.auth.decorators import login_required, user_passes_test
from django.forms import ValidationError
from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from django.http import HttpResponse
import datetime as dt
import io
from datetime import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.db.models import Q
import numpy as np
from django.db.models import Avg, Sum
from django.utils.html import escape
# Proyect imports
from .models import *


# Create your views here.
@login_required
def home(request):
    return render(request, "home.html", {
        "user": request.user
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas']).exists())
def cotizacion_menu(request):
    if request.method == "GET":
        return render(request, "cotizacion_menu.html")
    if request.method == "POST":
        tipo_cotizacion = request.POST["tipo_cotizacion"]
        print("## TIPO COTIZACION ##: ", tipo_cotizacion)
        return redirect("cotizacion_form", tipo_cotizacion)

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas']).exists())
def cotizacion_form(request, tipo_cotizacion):
    if request.method == "GET":
        vidrio_list = Articulos.objects.filter(tipo="VIDRIO").only("articulo")
        tinta_vit_list = Articulos.objects.filter(tipo="TINTA_VITRIFICABLE").only("articulo")
        tinta_UV_list = Articulos.objects.filter(tipo="TINTA_UV").only("articulo")
        tinta_epox_list = Articulos.objects.filter(tipo="TINTA_EPOXI").only("articulo")
        pintura_sopleteado_list = Articulos.objects.filter(tipo="PINTURA_SOPLETEADO").only("articulo")
        ctg_tipo_list = SopleteadoParametros.objects.all()
        etiquetas_list = Articulos.objects.filter(tipo="ETIQUETA").only("articulo")
        embalaje_list = Articulos.objects.filter(tipo="EMBALAJE_ESPECIAL").only("articulo")
        clientes_list = Clientes.objects.all()
        proyectos_list = CosteoSerigrafia.objects.exclude(proyecto__isnull=True).values('proyecto').distinct()
        print("proyectos_list: ", proyectos_list)
        iibb_bsas = FinanzasParametros.objects.filter(id=1).values()[0]["iibb_bsas"]
        iibb_mza = FinanzasParametros.objects.filter(id=1).values()[0]["iibb_mza"]
        hig_seg = FinanzasParametros.objects.filter(id=1).values()[0]["hig_seg"]

        return render(request, "cotizacion_form.html", {
            "tipo_cotizacion": tipo_cotizacion,
            "vidrio_list":vidrio_list,
            "tinta_vit_list":tinta_vit_list,
            "tinta_UV_list":tinta_UV_list,
            "tinta_epox_list":tinta_epox_list,
            "etiquetas_list":etiquetas_list,
            "embalaje_list":embalaje_list,
            "clientes_list":clientes_list,
            "proyectos_list":proyectos_list,
            "pintura_sopleteado_list":pintura_sopleteado_list,
            "ctg_tipo_list":ctg_tipo_list,
            "iibb_bsas":iibb_bsas,
            "iibb_mza":iibb_mza,
            "hig_seg":hig_seg
        })
    
    if request.method == "POST":
        # 1. Convierto str a booleanos
        # print("## POST PRE CAMBIOS ##: ", request.POST)
        request.POST._mutable = True
        keys_list = list(request.POST.keys())
        for item in keys_list:
            if request.POST[item] == "True" or request.POST[item] == "False":
                if request.POST[item] == "True":
                    request.POST[item] = True
                else:
                    request.POST[item] = False
            elif request.POST[item] == "":
                request.POST[item] = None
        request.POST["created_by"] = request.user
        request.POST["tipo_cotiz"] = tipo_cotizacion
        try:
            request.POST["cliente"] = Clientes.objects.get(cliente=request.POST["cliente"])
        except Clientes.DoesNotExist:
            # El cliente no está en la base de datos. El error lo mostrará la validación del formulario, por eso "pass".
            pass
        request.POST._mutable = False
        # print("## POST POST CAMBIOS ##: ",request.POST)               
        # 2. Valido formulario
        form = CosteoSerigrafiaForm(request.POST)
        if form.is_valid():
            try:
                articulo = Articulos.objects.get(articulo=request.POST["articulo"])
            except Articulos.DoesNotExist:
                form.add_error("articulo", ValidationError("El artículo seleccionado no existe"))
                messages.warning(request, f"Envío de datos fallido: {form.errors}")
                return redirect("cotizacion_menu")
            # guardar el "Articulo.tipo_2" en una variable
            tipo_2 = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["tipo_2"]
            velocidad = Decimal(request.POST["velocidad"])
            # 3. Calculos de costeo y guardado en CotizadorSerigrafia
            if "Serigrafia" in tipo_cotizacion:
                vel_tej_1 = ProduccionParametros.objects.filter(id=1).values()[0]["horno_vel_tejido_1"]
                vel_tej_2 = ProduccionParametros.objects.filter(id=1).values()[0]["horno_vel_tejido_2"]
                cant_cinta_1 = Decimal(request.POST["horno"])
                cant_cinta_2 = Decimal(request.POST["horno"])
                vel_horno = ((cant_cinta_1+cant_cinta_2)**Decimal(2))/Decimal(1.8)*(vel_tej_1+vel_tej_2)/Decimal(200)*Decimal(60)
                persmaq_horno = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_horno"]
                cant_colores = 0
                if request.POST["cant_colores"] != None:
                    cant_colores = int(request.POST["cant_colores"])

            costo_hora_operario = RecursosHumanos.objects.filter(puesto='OPERARIO').aggregate(promedio=Avg('costo_hora_total'))['promedio']
            costo_hora_maquinista = RecursosHumanos.objects.filter(puesto='MAQUINISTA').aggregate(promedio=Avg('costo_hora_total'))['promedio']
            costo_hora_laboratorio = RecursosHumanos.objects.filter(puesto='LABORATORIO').aggregate(promedio=Avg('costo_hora_total'))['promedio']
            costo_hora_autoelevador = RecursosHumanos.objects.filter(puesto='AUTOELEVADOR').aggregate(promedio=Avg('costo_hora_total'))['promedio']
            costo_hora_logistica = RecursosHumanos.objects.filter(puesto='LOGISTICA').aggregate(promedio=Avg('costo_hora_total'))['promedio']
            costo_mensual_moind = RecursosHumanos.objects.filter(Q(puesto="LOGISTICA") | Q(puesto="MANTENIMIENTO")).aggregate(suma=Sum('costo_mensual_total'))["suma"]
            costo_mensual_logistica = RecursosHumanos.objects.filter(puesto='LOGISTICA').aggregate(suma=Sum('costo_mensual_total'))["suma"]
            costo_mensual_mantenimiento = RecursosHumanos.objects.filter(puesto='MANTENIMIENTO').aggregate(suma=Sum('costo_mensual_total'))["suma"]
            costo_mensual_disenio = RecursosHumanos.objects.filter(puesto='DISEÑO').aggregate(suma=Sum('costo_mensual_total'))["suma"]
            gas = Articulos.objects.filter(articulo='Gas').values("costo_ars")[0]["costo_ars"]
            luz = Articulos.objects.filter(articulo='Luz').values("costo_ars")[0]["costo_ars"]
            gral_prod_mens_bsas = ProduccionParametros.objects.filter(id=1).values()[0]["gral_prod_mens_bsas"]
            gral_prod_mens_mza = ProduccionParametros.objects.filter(id=1).values()[0]["gral_prod_mens_mza"]
            cant_unidades = Decimal(request.POST["cant_unidades"])
            screen_cant_art = ProduccionParametros.objects.filter(id=1).values()[0]["screen_cant_art"]
            aux = Decimal(ceil(cant_unidades/screen_cant_art)/cant_unidades)
            gral_malos_color_lavadero = ProduccionParametros.objects.filter(id=1).values()[0]["gral_malos_color_lavadero"]/100
            costo_deuda = FinanzasParametros.objects.filter(id=1).values()[0]["costo_deuda"]/100
            persmaq_autoelevador = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_autoelevador"]
            ins_unid_film = ProduccionParametros.objects.filter(id=1).values()[0]["ins_unid_film"]
            ins_unid_cinta = ProduccionParametros.objects.filter(id=1).values()[0]["ins_unid_cinta"]
            costo_film_stretch = Articulos.objects.filter(articulo='Film stretch (12 kilos por rollo, 90 rollos por pallet) U$D 2,85 x kilo').values("costo_ars")[0]["costo_ars"]
            costo_cinta = Articulos.objects.filter(articulo='Cinta').values("costo_ars")[0]["costo_ars"]
            var = Decimal(0)
            rot_horno_propia = 0
            rot_impresion_propia = 0
            # CAMPOS COTIZACION SERIGRAFIA:#
            # GENERALES
            value_mp_articulo = 0
            value_codigo_caja = 0
            value_embalaje = 0
            value_transporte = 0
            # SERIGRAFIA
            value_quemado = 0
            value_hh_cambio_screen = 0
            value_hh_decoracion = 0
            value_hh_horno = 0
            value_mp_screen = 0
            value_hh_screen_cop = 0
            value_luz_gas = 0
            value_mp_pintura = 0
            value_sampista = 0
            value_rot_impresion = 0
            value_rot_horno = 0
            value_calco_vitrificable = 0
            value_mp_cinta_stretch = 0
            value_logistica = 0
            value_setup = 0
            value_disenio_pelicula = 0
            value_lavado = 0
            value_mantenimiento = 0
            # COATING
            value_ctg_setup = 0
            value_ctg_hh_sopleteadora = 0
            value_ctg_luz_gas = 0
            value_ctg_mp_pintura = 0
            value_ctg_sampista = 0
            value_ctg_rotura = 0
            value_ctg_lavado = 0
            value_ctg_mp_cinta_stretch = 0
            value_ctg_mo_indirecta = 0

        # GENERALES 1/2 - Campos #
            # 3.1 mp_articulo
            if request.POST["art_propio"]:
                # Si es articulo propio (True)
                costo_mp_articulo = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["costo_ars"]
                value_mp_articulo = (costo_mp_articulo)*cant_unidades
            # 3.11 codigo_caja
            if request.POST["codigo"] != None:
                costo_codigo_caja = Articulos.objects.filter(articulo=request.POST["codigo"]).values("costo_ars")[0]["costo_ars"]
                value_codigo_caja = (costo_codigo_caja+1/velocidad*costo_hora_operario)*cant_unidades
            # 3.12 costo_embalaje_esp
            if request.POST["embalaje"] != None:
                costo_embalaje_esp = Articulos.objects.filter(articulo=request.POST["embalaje"]).values("costo_ars")[0]["costo_ars"]
                velocidad_aplicacion_u_h_embalaje_esp = Articulos.objects.filter(articulo=request.POST["embalaje"]).values("velocidad_aplicacion_u_h")[0]["velocidad_aplicacion_u_h"]
                value_embalaje = (costo_embalaje_esp+1/velocidad_aplicacion_u_h_embalaje_esp*costo_hora_operario)*cant_unidades
            # 3.16 transporte
            if request.POST["transporte"]:
                value_transporte = Decimal(request.POST["transporte"])

        # SERIGRAFIA #
            if "Serigrafia" in tipo_cotizacion:
                # 3.2 quemado_etiq
                if request.POST["quemado"]:
                    # Si hay quemado (True)
                    value_quemado = (1/vel_horno*persmaq_horno*costo_hora_operario+(gas/gral_prod_mens_bsas))*cant_unidades
                # 3.3 hh_cambio_screen, hh_decoracion, mp_screen, hh_screen_cop, calco_vitrificable, setup
                if request.POST["tipo_decorado"] == "Calco Vitrificable":
                    costo_calco = Decimal(request.POST["costo_calco"])
                    persmaq_vel_calco = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_vel_calco"]
                    value_calco_vitrificable = (costo_calco+1/persmaq_vel_calco*costo_hora_operario)*cant_unidades
                else:
                # Si no es "Calco Vitrificable" (osea, si es tinta)
                    # hh_cambio_screen
                    screen_tpo_cambio = ProduccionParametros.objects.filter(id=1).values()[0]["screen_tpo_cambio"]
                    value_hh_cambio_screen = (aux*screen_tpo_cambio*costo_hora_maquinista*cant_colores)*cant_unidades
                    # hh_decoracion
                    cant_personal = Decimal(request.POST["cant_personal"])
                    persmaq_personal_asa = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_personal_asa"]
                    asa = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["asa"]
                    if asa == "SI":
                        flag_asa = 1
                    else:
                        flag_asa = 0
                    persmaq_maquinistas = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_maquinistas"]
                    value_hh_decoracion = ((cant_personal+persmaq_personal_asa*cant_colores*flag_asa)/velocidad*costo_hora_operario+persmaq_maquinistas/velocidad*costo_hora_maquinista)*cant_unidades
                    # mp_screen
                    screen_medida = ProduccionParametros.objects.filter(id=1).values()[0]["screen_medida"]
                    costo_polyester_120 = Articulos.objects.filter(articulo='Polyester 120').values("costo_ars")[0]["costo_ars"]
                    value_mp_screen = (2*screen_medida/10000*costo_polyester_120*Decimal(aux)*cant_colores)*cant_unidades
                    # hh_screen_cop
                    screen_fab_hora = ProduccionParametros.objects.filter(id=1).values()[0]["screen_fab_hora"]
                    value_hh_screen_cop = (Decimal(aux)*screen_fab_hora*costo_hora_laboratorio*cant_colores)*cant_unidades
                    if (request.POST["maquina_auto"]):
                        var = Decimal(1)
                    else:
                        var = Decimal(0.15)
                    value_setup = var*Decimal(2+cant_colores)*Decimal(2)*costo_hora_maquinista
                # 3.4 hh_horno
                if (request.POST["tipo_decorado"] != "Tinta UV") and (request.POST["tipo_decorado"] != "Tinta epoxi"):
                    value_hh_horno = (Decimal(1)/vel_horno*persmaq_horno*costo_hora_operario)*cant_unidades
                # 3.5 luz_gas
                value_luz_gas = ((gas+luz)/gral_prod_mens_bsas)*cant_unidades
                # 3.6 mp_pintura
                if cant_colores:
                    for n in range(1, cant_colores+1):
                        value_dibujo = Decimal(request.POST[f"dibujo_{n}"])
                        name_pintura = request.POST[f"pintura_{n}"]
                        costo_pintura = Articulos.objects.filter(articulo=name_pintura).values()[0]["costo_ars"]/1000
                        consumo_pintura = Articulos.objects.filter(articulo=name_pintura).values()[0]["consumo_g_cm2"]
                        value_relieve = Decimal(request.POST[f"relieve_{n}"])
                        costo_pintura_n = value_dibujo*costo_pintura*value_relieve*consumo_pintura
                        value_mp_pintura = value_mp_pintura+costo_pintura_n
                    value_mp_pintura = value_mp_pintura*cant_unidades
                # 3.7 sampista
                value_sampista = (1/velocidad*costo_hora_autoelevador*persmaq_autoelevador)*cant_unidades
                # 3.8 rot_impresion, rot_horno
                gral_rotura_maquina = ProduccionParametros.objects.filter(id=1).values()[0]["gral_rotura_maquina"]/100
                rot_impresion_propia = gral_rotura_maquina*value_mp_articulo
                rot_impresion_cliente = gral_malos_color_lavadero*Decimal(value_mp_pintura)
                value_rot_impresion = rot_impresion_propia+rot_impresion_cliente
                gral_rotura_horno = ProduccionParametros.objects.filter(id=1).values()[0]["gral_rotura_horno"]/100
                rot_horno_propia = gral_rotura_horno*value_mp_articulo
                rot_horno_cliente = gral_rotura_horno*Decimal(value_mp_pintura)
                value_rot_horno = rot_horno_propia+rot_horno_cliente
                # 3.9 lavado
                lav_limp_ser = ProduccionParametros.objects.filter(id=1).values()[0]["lav_limp_ser"]
                value_lavado = (Decimal(1)/lav_limp_ser*costo_hora_operario*gral_malos_color_lavadero)*cant_unidades
                # 3.10 mp_cinta_stretch
                value_mp_cinta_stretch = (Decimal(1)/ins_unid_film*costo_film_stretch+Decimal(1)/ins_unid_cinta*costo_cinta)*cant_unidades
                # 3.13 logistica
                value_logistica = (costo_mensual_logistica/gral_prod_mens_bsas)*cant_unidades
                # 3.14 mantenimiento
                value_mantenimiento = (costo_mensual_mantenimiento/gral_prod_mens_bsas)*cant_unidades
                # 3.15 disenio_pelicula
                if request.POST["disenio"] and (cant_colores != None):
                    disenio_perc = Decimal(request.POST["disenio_perc"])/100
                    costo_x_pelicula = Articulos.objects.filter(articulo=request.POST["articulo"]).values("costo_ars")[0]["costo_ars"]
                    dis_cant_bsas = ProduccionParametros.objects.filter(id=1).values()[0]["dis_cant_bsas"]
                    dis_cant_mza = ProduccionParametros.objects.filter(id=1).values()[0]["dis_cant_mza"]
                    value_disenio_pelicula = disenio_perc*(costo_mensual_disenio/(dis_cant_bsas+dis_cant_mza)+costo_x_pelicula)*cant_colores
        # COATING #
            if "Coating" in tipo_cotizacion:
                # ctg_setup
                sopl_operarios = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_operarios"]
                sopl_tpo_setup = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_tpo_setup"]
                value_ctg_setup = sopl_operarios*sopl_tpo_setup*costo_hora_operario
                # ctg_hh_sopleteadora
                ctg_cant_pasadas = Decimal(request.POST["ctg_cant_pasadas"])
                ctg_tipo = request.POST["ctg_tipo"]
                tipo_velocidad = SopleteadoParametros.objects.filter(tipo_articulo=ctg_tipo).values()[0]["velocidad"]
                sopl_maquinistas = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_maquinistas"]
                value_ctg_hh_sopleteadora = ctg_cant_pasadas*(sopl_operarios*cant_unidades/tipo_velocidad*costo_hora_operario+sopl_maquinistas*cant_unidades/tipo_velocidad*costo_hora_maquinista)
                # guardo en el POST el "tipo"
                # request.POST._mutable = True
                # request.POST["ctg_tipo"] = SopleteadoParametros.objects.get(tipo_articulo=request.POST["ctg_tipo"])
                # request.POST._mutable = False
                # print("## POST POST CAMBIOS ##: ",request.POST)
                # ctg_luz_gas
                value_ctg_luz_gas = ctg_cant_pasadas*(cant_unidades*(luz+gas)/gral_prod_mens_bsas)
                # ctg_mp_pintura
                costo_g_solvente = Articulos.objects.filter(articulo="L-DSOP-0011 Solvente X LITRO").values("costo_ars")[0]["costo_ars"]/Decimal(1000)
                costo_g_aditivo = Articulos.objects.filter(articulo="900.99.0125 Aditivo X KG").values("costo_ars")[0]["costo_ars"]/Decimal(1000)
                consumo_solv_adit_pint = SopleteadoParametros.objects.filter(tipo_articulo=ctg_tipo).values()[0]["consumo"]
                sopl_perc_solv = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_perc_solv"]/Decimal(100)
                sopl_perc_adit = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_perc_adit"]/Decimal(100)
                sopl_perc_pint = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_perc_pint"]/Decimal(100)
                cant_g_solv = consumo_solv_adit_pint*sopl_perc_solv
                cant_g_adit = consumo_solv_adit_pint*sopl_perc_adit
                cant_g_pintura = consumo_solv_adit_pint*sopl_perc_pint
                costo_unit_solv = costo_g_solvente*cant_g_solv
                costo_unit_adit = costo_g_aditivo*cant_g_adit
                ctg_cant_colores = int(request.POST["ctg_cant_colores"])
                rotura = Decimal(1.1)
                for n in range(1, ctg_cant_colores+1):
                    name_pintura = request.POST[f"ctg_color_{n}"]
                    ctg_cantidad_n = Decimal(request.POST[f"ctg_cantidad_{n}"])
                    ctg_cantidad_n = ctg_cantidad_n*rotura # Agrego roturas 10 %
                    costo_g_pintura = Articulos.objects.filter(articulo=name_pintura).values("costo_ars")[0]["costo_ars"]/Decimal(1000)
                    costo_unit_pint = costo_g_pintura*cant_g_pintura
                    costo_unit_color = costo_unit_pint+costo_unit_solv+costo_unit_adit
                    value_ctg_mp_pintura += costo_unit_color*ctg_cantidad_n        
                # ctg_sampista
                value_ctg_sampista = ctg_cant_pasadas*(cant_unidades/tipo_velocidad*costo_hora_autoelevador*persmaq_autoelevador)
                # ctg_rotura
                sopl_rotura = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_rotura"]/Decimal(100)
                value_ctg_rotura = ctg_cant_pasadas*(value_mp_articulo*sopl_rotura)
                # ctg_lavado
                lav_limp_sopl = ProduccionParametros.objects.filter(id=1).values()[0]["lav_limp_sopl"]
                sopl_lavadero = ProduccionParametros.objects.filter(id=1).values()[0]["sopl_lavadero"]/Decimal(100)
                value_ctg_lavado = ctg_cant_pasadas*(cant_unidades/lav_limp_sopl*sopl_lavadero*costo_hora_operario)
                # ctg_mp_cinta_stretch
                value_ctg_mp_cinta_stretch = (Decimal(1)/ins_unid_film*costo_film_stretch+Decimal(1)/ins_unid_cinta*costo_cinta)*cant_unidades
                # ctg_mo_indirecta
                value_ctg_mo_indirecta = costo_mensual_moind*cant_unidades/gral_prod_mens_bsas

        # CÁLCULO DE COSTOS #
            # SERIGRAFIA #

            # COATING #

            # TOTALES #
            # 3.17 costo_produccion
            value_costo_produccion = value_quemado+value_hh_cambio_screen+value_hh_decoracion+ \
                                        value_hh_horno+value_mp_screen+value_hh_screen_cop+value_luz_gas+value_mp_pintura+ \
                                        value_sampista+value_rot_impresion+value_rot_horno+value_lavado+value_calco_vitrificable+ \
                                        value_mp_cinta_stretch+value_logistica+value_mantenimiento+value_setup+ \
                                        value_disenio_pelicula+\
                                        value_ctg_setup+value_ctg_hh_sopleteadora+value_ctg_luz_gas+value_ctg_mp_pintura+\
                                        value_ctg_sampista+value_ctg_rotura+value_ctg_lavado+value_ctg_mp_cinta_stretch+\
                                        value_ctg_mo_indirecta+\
                                        value_codigo_caja+value_transporte+value_mp_articulo+value_embalaje \
            # 3.18 costo_mp
            value_costo_mp = value_mp_articulo+value_mp_cinta_stretch+value_mp_pintura+value_mp_screen+\
                                value_ctg_mp_pintura+value_ctg_mp_cinta_stretch
            # 3.19 costo_funcionamiento
            finan_costo_funcionamiento = FinanzasParametros.objects.filter(id=1).values()[0]["costo_funcionamiento"]
            perc_articulo = FinanzasParametros.objects.filter(id=1).values()[0]["perc_propio"]/Decimal(100)
            perc_decorado = FinanzasParametros.objects.filter(id=1).values()[0]["perc_cliente"]/Decimal(100)
            ventas_estimadas_articulo = FinanzasParametros.objects.filter(id=1).values()[0]["ventas_estimadas_propio"]
            ventas_estimadas_decorado = FinanzasParametros.objects.filter(id=1).values()[0]["ventas_estimadas_cliente"]
            # ventas_estimadas_propio_art = FinanzasParametros.objects.filter(id=1).values()[0]["compras_estimadas_propio"]
            # ventas_estimadas_propio_cliente = ventas_estimadas_propio-ventas_estimadas_propio_art
            # ventas_estimadas_deco_total = ventas_estimadas_propio_cliente+ventas_estimadas_cliente
            perc_costo_decorado = (perc_decorado*finan_costo_funcionamiento)/ventas_estimadas_decorado
            perc_costo_articulo = (perc_articulo*finan_costo_funcionamiento)/ventas_estimadas_articulo
            costos_articulo = value_mp_articulo+rot_horno_propia+rot_impresion_propia+value_ctg_rotura
            costos_decorado = value_costo_produccion-costos_articulo
            value_costo_funcionamiento_articulo = costos_articulo*(perc_costo_articulo/(1-perc_costo_articulo))
            value_costo_funcionamiento_decorado = costos_decorado*(perc_costo_decorado/(1-perc_costo_decorado))
            print("## CALCULO PRESUPUESTO VENTAS ##")
            print("ventas_estimadas_articulo: ", ventas_estimadas_articulo)
            print("ventas_estimadas_decorado: ", ventas_estimadas_decorado)
            # print("ventas_estimadas_propio_art: ", ventas_estimadas_propio_art)
            # print("ventas_estimadas_propio_cliente: ", ventas_estimadas_propio_cliente)
            # print("ventas_estimadas_deco_total: ", ventas_estimadas_deco_total)

            print("\n## CALCULO COSTOS ##")
            print("perc_decorado: ", perc_decorado)
            print("perc_articulo: ", perc_articulo)                      
            print("value_costo_funcionamiento_decorado: ", value_costo_funcionamiento_decorado)
            print("value_costo_funcionamiento_articulo: ", value_costo_funcionamiento_articulo)
            print("costos_decorado: ", costos_decorado)
            print("costos_articulo: ", costos_articulo)
            print("perc_costo_decorado: ", perc_costo_decorado)
            print("perc_costo_articulo: ", perc_costo_articulo)
            print("perc_costo_decorado/(1-perc_costo_decorado): ", perc_costo_decorado/(1-perc_costo_decorado))
            print("perc_costo_articulo/(1-perc_costo_articulo): ", perc_costo_articulo/(1-perc_costo_articulo))

            # print("ventas_estimadas_deco_total: ", ventas_estimadas_deco_total)
            value_costo_funcionamiento_total = value_costo_funcionamiento_articulo+value_costo_funcionamiento_decorado
            # 3.20 costo_total
            value_costo_total = value_costo_produccion+value_costo_funcionamiento_total
            # 3.21 costo_unitario fabricacion
            value_costo_unitario_fab = value_costo_total/cant_unidades
            value_costo_unitario_fab_articulo = (costos_articulo+value_costo_funcionamiento_articulo)/cant_unidades
            value_costo_unitario_fab_decorado = (costos_decorado+value_costo_funcionamiento_decorado)/cant_unidades
            # 3.22 costos con ajustes financieros e impositivos
            value_costo_unitario = value_costo_unitario_fab
            list_cargos = ["iibb_bsas", "iibb_mza", "hig_seg", "com_ventas", "com_otras"]
            for cargo in list_cargos:
                if request.POST[cargo]:
                    value_costo_unitario += value_costo_unitario_fab/Decimal(1-Decimal(request.POST[cargo])/100)-value_costo_unitario_fab
            value_costo_unitario_15 = (value_costo_unitario)*(1+costo_deuda/Decimal(2)) # Calculado para 15 días
            value_costo_unitario_30 = (value_costo_unitario)*(1+costo_deuda) # Calculado para 30 días
            value_costo_unitario_45 = (value_costo_unitario)*(1+costo_deuda*Decimal(1.5)) # Calculado para 45 días
            value_costo_unitario_60 = (value_costo_unitario)*(1+costo_deuda*Decimal(2)) # Calculado para 60 días

            # 4. Ya realizados los cálculos, si no saltó error guardo los datos en CosteoSerigrafia (tanto serigrafía como coating)
            new_costeo_serig = form.save(commit=False)
            for field_name, field_value in form.cleaned_data.items():
                setattr(new_costeo_serig, field_name, field_value)
            new_costeo_serig.tipo_2 = tipo_2
            new_costeo_serig.save()

            # Luego de guardar lo de CosteoSerigrafia, se guardan los datos en la tabla CotizadorSerigrafia
            new_cotizador = CotizadorSerigrafia(srg_quemado=value_quemado, srg_hh_cambio_screen=value_hh_cambio_screen, \
                                                srg_hh_decoracion=value_hh_decoracion, srg_hh_horno=value_hh_horno, \
                                                srg_mp_screen=value_mp_screen, srg_hh_screen_cop=value_hh_screen_cop, \
                                                srg_luz_gas=value_luz_gas, srg_mp_pintura=value_mp_pintura, srg_sampista=value_sampista, \
                                                srg_rot_impresion=value_rot_impresion, srg_rot_horno=value_rot_horno, srg_lavado=value_lavado, \
                                                srg_calco_vitrificable=value_calco_vitrificable, srg_mp_cinta_stretch=value_mp_cinta_stretch, \
                                                srg_logistica=value_logistica, srg_mantenimiento=value_mantenimiento, srg_setup=value_setup, \
                                                srg_disenio_pelicula=value_disenio_pelicula, \
                                                ctg_setup=value_ctg_setup, ctg_hh_sopleteadora=value_ctg_hh_sopleteadora, ctg_luz_gas=value_ctg_luz_gas, \
                                                ctg_mp_pintura=value_ctg_mp_pintura, ctg_sampista=value_ctg_sampista, ctg_rotura=value_ctg_rotura, \
                                                ctg_lavado=value_ctg_lavado, ctg_mp_cinta_stretch=value_ctg_mp_cinta_stretch, ctg_mo_indirecta=value_ctg_mo_indirecta, \
                                                mp_articulo=value_mp_articulo, codigo_caja=value_codigo_caja, embalaje=value_embalaje, transporte=value_transporte, \
                                                costo_unitario=value_costo_unitario, costo_unitario_15=value_costo_unitario_15, \
                                                costo_unitario_30=value_costo_unitario_30, costo_unitario_45=value_costo_unitario_45, \
                                                costo_unitario_60=value_costo_unitario_60, costo_produccion=value_costo_produccion, \
                                                costo_mp=value_costo_mp, costo_funcionamiento_articulo = value_costo_funcionamiento_articulo, \
                                                costo_funcionamiento_decorado = value_costo_funcionamiento_decorado, \
                                                costo_funcionamiento_total=value_costo_funcionamiento_total, costo_total=value_costo_total,\
                                                costo_unitario_fab=value_costo_unitario_fab, created_by=request.user, \
                                                costo_unitario_fab_articulo=value_costo_unitario_fab_articulo, costo_unitario_fab_decorado=value_costo_unitario_fab_decorado)

            new_cotizador.save()
            new_costeo_serig.cotizacion = new_cotizador
            new_costeo_serig.save()
            messages.success(request, "Envío de datos exitoso")
        else:
            messages.warning(request, f"Envío de datos fallido: {form.errors}")
            return redirect("cotizacion_menu")
            
        return redirect("cotizaciones")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas']).exists())
def costeo_serigrafia(request):
    if request.method == "GET":
        vidrio_list = Articulos.objects.filter(tipo="VIDRIO").only("articulo")
        tinta_vit_list = Articulos.objects.filter(tipo="TINTA_VITRIFICABLE").only("articulo")
        tinta_UV_list = Articulos.objects.filter(tipo="TINTA_UV").only("articulo")
        tinta_epox_list = Articulos.objects.filter(tipo="TINTA_EPOXI").only("articulo")
        etiquetas_list = Articulos.objects.filter(tipo="ETIQUETA").only("articulo")
        embalaje_list = Articulos.objects.filter(tipo="EMBALAJE_ESPECIAL").only("articulo")
        clientes_list = Clientes.objects.all()
        iibb_bsas = FinanzasParametros.objects.filter(id=1).values()[0]["iibb_bsas"]
        iibb_mza = FinanzasParametros.objects.filter(id=1).values()[0]["iibb_mza"]
        hig_seg = FinanzasParametros.objects.filter(id=1).values()[0]["hig_seg"]

        return render(request, "costeo_serigrafia.html", {
            "vidrio_list":vidrio_list,
            "tinta_vit_list":tinta_vit_list,
            "tinta_UV_list":tinta_UV_list,
            "tinta_epox_list":tinta_epox_list,
            "etiquetas_list":etiquetas_list,
            "embalaje_list":embalaje_list,
            "clientes_list":clientes_list,
            "iibb_bsas":iibb_bsas,
            "iibb_mza":iibb_mza,
            "hig_seg":hig_seg
        })
    
    # if request.method == "POST":
    #     # 1. Convierto str a booleanos
    #     # print("## POST PRE CAMBIOS ##: ", request.POST)
    #     request.POST._mutable = True
    #     keys_list = list(request.POST.keys())
    #     for item in keys_list:
    #         if request.POST[item] == "True" or request.POST[item] == "False":
    #             if request.POST[item] == "True":
    #                 request.POST[item] = True
    #             else:
    #                 request.POST[item] = False
    #         elif request.POST[item] == "":
    #             request.POST[item] = None
    #     request.POST["created_by"] = request.user

    #     try:
    #         request.POST["cliente"] = Clientes.objects.get(cliente=request.POST["cliente"])
    #     except Clientes.DoesNotExist:
    #         # El cliente no está en la base de datos. El error lo mostrará la validación del formulario, por eso "pass".
    #         pass
    #     request.POST._mutable = False
    #     # print("## POST POST CAMBIOS ##: ",request.POST)
    #     # 2. Valido formulario
    #     form = CosteoSerigrafiaForm(request.POST)
    #     if form.is_valid():
    #         try:
    #             articulo = Articulos.objects.get(articulo=request.POST["articulo"])
    #         except Articulos.DoesNotExist:
    #             form.add_error("articulo", ValidationError("El artículo seleccionado no existe"))
    #             messages.warning(request, f"Envío de datos fallido: {form.errors}")
    #             return redirect("costeo_serigrafia")
    #         # guardar el "Articulo.tipo_2" en una variable
    #         tipo_2 = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["tipo_2"]
    #         # 3. Calculos de costeo y guardado en CotizadorSerigrafia
    #         # Grabo en variables los valores a usar en común en los diferentes cálculos, y también las variables q se asignaran luego como valor a los campos.
    #         # En comun:
    #         vel_tej_1 = ProduccionParametros.objects.filter(id=1).values()[0]["horno_vel_tejido_1"]
    #         vel_tej_2 = ProduccionParametros.objects.filter(id=1).values()[0]["horno_vel_tejido_2"]
    #         cant_cinta_1 = Decimal(request.POST["horno"])
    #         cant_cinta_2 = Decimal(request.POST["horno"])
    #         vel_horno = ((cant_cinta_1+cant_cinta_2)**Decimal(2))/Decimal(1.8)*(vel_tej_1+vel_tej_2)/Decimal(200)*Decimal(60)
    #         persmaq_horno = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_horno"]
    #         costo_hora_operario = RecursosHumanos.objects.filter(puesto='OPERARIO').aggregate(promedio=Avg('costo_hora_total'))['promedio']
    #         costo_hora_maquinista = RecursosHumanos.objects.filter(puesto='MAQUINISTA').aggregate(promedio=Avg('costo_hora_total'))['promedio']
    #         costo_hora_laboratorio = RecursosHumanos.objects.filter(puesto='LABORATORIO').aggregate(promedio=Avg('costo_hora_total'))['promedio']
    #         costo_hora_autoelevador = RecursosHumanos.objects.filter(puesto='AUTOELEVADOR').aggregate(promedio=Avg('costo_hora_total'))['promedio']
    #         costo_hora_logistica = RecursosHumanos.objects.filter(puesto='LOGISTICA').aggregate(promedio=Avg('costo_hora_total'))['promedio']
    #         costo_mensual_logistica = RecursosHumanos.objects.filter(puesto='LOGISTICA').aggregate(suma=Sum('costo_mensual_total'))["suma"]
    #         costo_mensual_mantenimiento = RecursosHumanos.objects.filter(puesto='MANTENIMIENTO').aggregate(suma=Sum('costo_mensual_total'))["suma"]
    #         costo_mensual_disenio = RecursosHumanos.objects.filter(puesto='DISEÑO').aggregate(suma=Sum('costo_mensual_total'))["suma"]
    #         gas = Articulos.objects.filter(articulo='Gas').values("costo_ars")[0]["costo_ars"]
    #         luz = Articulos.objects.filter(articulo='Luz').values("costo_ars")[0]["costo_ars"]
    #         gral_prod_mens_bsas = ProduccionParametros.objects.filter(id=1).values()[0]["gral_prod_mens_bsas"]
    #         gral_prod_mens_mza = ProduccionParametros.objects.filter(id=1).values()[0]["gral_prod_mens_mza"]
    #         cant_unidades = Decimal(request.POST["cant_unidades"])
    #         screen_cant_art = ProduccionParametros.objects.filter(id=1).values()[0]["screen_cant_art"]
    #         aux = Decimal(ceil(cant_unidades/screen_cant_art)/cant_unidades)
    #         velocidad = Decimal(request.POST["velocidad"])
    #         gral_malos_color_lavadero = ProduccionParametros.objects.filter(id=1).values()[0]["gral_malos_color_lavadero"]/100
    #         costo_deuda = FinanzasParametros.objects.filter(id=1).values()[0]["costo_deuda"]/100
    #         persmaq_autoelevador = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_autoelevador"]
    #         cant_colores = 0
    #         if request.POST["cant_colores"] != None:
    #             cant_colores = int(request.POST["cant_colores"])
    #         var = Decimal(0)
    #         # Campos:
    #         value_mp_articulo = 0
    #         value_quemado = 0
    #         value_hh_cambio_screen = 0
    #         value_hh_decoracion = 0
    #         value_hh_horno = 0
    #         value_mp_screen = 0
    #         value_hh_screen_cop = 0
    #         value_luz_gas = 0
    #         value_mp_pintura = 0
    #         value_sampista = 0
    #         value_rot_impresion = 0
    #         value_rot_horno = 0
    #         value_calco_vitrificable = 0
    #         value_mp_cinta_stretch = 0
    #         value_codigo_caja = 0
    #         value_embalaje = 0
    #         value_logistica = 0
    #         value_setup = 0
    #         value_disenio_pelicula = 0
    #         value_transporte = 0
            
    #         # 3.1 mp_articulo
    #         if request.POST["art_propio"]:
    #             # Si es articulo propio (True)
    #             costo_mp_articulo = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["costo_ars"]
    #             value_mp_articulo = (costo_mp_articulo)*cant_unidades
    #         # 3.2 quemado_etiq
    #         if request.POST["quemado"]:
    #             # Si hay quemado (True)
    #             value_quemado = (1/vel_horno*persmaq_horno*costo_hora_operario+(gas/gral_prod_mens_bsas))*cant_unidades
    #         # 3.3 hh_cambio_screen, hh_decoracion, mp_screen, hh_screen_cop, calco_vitrificable, setup
    #         if request.POST["tipo_decorado"] == "Calco Vitrificable":
    #             costo_calco = Decimal(request.POST["costo_calco"])
    #             persmaq_vel_calco = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_vel_calco"]
    #             value_calco_vitrificable = (costo_calco+1/persmaq_vel_calco*costo_hora_operario)*cant_unidades
    #         else:
    #         # Si no es "Calco Vitrificable" (osea, si es tinta)
    #             # hh_cambio_screen
    #             screen_tpo_cambio = ProduccionParametros.objects.filter(id=1).values()[0]["screen_tpo_cambio"]
    #             value_hh_cambio_screen = (aux*screen_tpo_cambio*costo_hora_maquinista*cant_colores)*cant_unidades
    #             # hh_decoracion
    #             cant_personal = Decimal(request.POST["cant_personal"])
    #             persmaq_personal_asa = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_personal_asa"]
    #             asa = Articulos.objects.filter(articulo=request.POST["articulo"]).values()[0]["asa"]
    #             if asa == "SI":
    #                 flag_asa = 1
    #             else:
    #                 flag_asa = 0
    #             persmaq_maquinistas = ProduccionParametros.objects.filter(id=1).values()[0]["persmaq_maquinistas"]
    #             value_hh_decoracion = ((cant_personal+persmaq_personal_asa*cant_colores*flag_asa)/velocidad*costo_hora_operario+persmaq_maquinistas/velocidad*costo_hora_maquinista)*cant_unidades
    #             # mp_screen
    #             screen_medida = ProduccionParametros.objects.filter(id=1).values()[0]["screen_medida"]
    #             costo_polyester_120 = Articulos.objects.filter(articulo='Polyester 120').values("costo_ars")[0]["costo_ars"]
    #             value_mp_screen = (2*screen_medida/10000*costo_polyester_120*Decimal(aux)*cant_colores)*cant_unidades
    #             # hh_screen_cop
    #             screen_fab_hora = ProduccionParametros.objects.filter(id=1).values()[0]["screen_fab_hora"]
    #             value_hh_screen_cop = (Decimal(aux)*screen_fab_hora*costo_hora_laboratorio*cant_colores)*cant_unidades
    #             if (request.POST["maquina_auto"]):
    #                 var = Decimal(1)
    #             else:
    #                 var = Decimal(0.15)
    #             value_setup = var*Decimal(2+cant_colores)*Decimal(2)*costo_hora_maquinista
    #         # 3.4 hh_horno
    #         if (request.POST["tipo_decorado"] != "Tinta UV") and (request.POST["tipo_decorado"] != "Tinta epoxi"):
    #             value_hh_horno = (Decimal(1)/vel_horno*persmaq_horno*costo_hora_operario)*cant_unidades
    #         # 3.5 luz_gas
    #         value_luz_gas = ((gas+luz)/gral_prod_mens_bsas)*cant_unidades
    #         # 3.6 mp_pintura
    #         if cant_colores:
    #             for n in range(1, cant_colores+1):
    #                 value_dibujo = Decimal(request.POST[f"dibujo_{n}"])
    #                 name_pintura = request.POST[f"pintura_{n}"]
    #                 costo_pintura = Articulos.objects.filter(articulo=name_pintura).values()[0]["costo_ars"]/1000
    #                 consumo_pintura = Articulos.objects.filter(articulo=name_pintura).values()[0]["consumo_g_cm2"]
    #                 value_relieve = Decimal(request.POST[f"relieve_{n}"])
    #                 costo_pintura_n = value_dibujo*costo_pintura*value_relieve*consumo_pintura
    #                 value_mp_pintura = value_mp_pintura+costo_pintura_n
    #             value_mp_pintura = value_mp_pintura*cant_unidades

    #         # 3.7 sampista
    #         value_sampista = (1/velocidad*costo_hora_autoelevador*persmaq_autoelevador)*cant_unidades
    #         # 3.8 rot_impresion, rot_horno
    #         gral_rotura_maquina = ProduccionParametros.objects.filter(id=1).values()[0]["gral_rotura_maquina"]/100
    #         rot_impresion_propia = gral_rotura_maquina*value_mp_articulo
    #         rot_impresion_cliente = gral_malos_color_lavadero*Decimal(value_mp_pintura)
    #         value_rot_impresion = rot_impresion_propia+rot_impresion_cliente
    #         gral_rotura_horno = ProduccionParametros.objects.filter(id=1).values()[0]["gral_rotura_horno"]/100
    #         rot_horno_propia = gral_rotura_horno*value_mp_articulo
    #         rot_horno_cliente = gral_rotura_horno*Decimal(value_mp_pintura)
    #         value_rot_horno = rot_horno_propia+rot_horno_cliente
    #         # 3.9 lavado
    #         lav_limp_ser = ProduccionParametros.objects.filter(id=1).values()[0]["lav_limp_ser"]
    #         value_lavado = (Decimal(1)/lav_limp_ser*costo_hora_operario*gral_malos_color_lavadero)*cant_unidades
    #         # 3.10 mp_cinta_stretch
    #         ins_unid_film = ProduccionParametros.objects.filter(id=1).values()[0]["ins_unid_film"]
    #         ins_unid_cinta = ProduccionParametros.objects.filter(id=1).values()[0]["ins_unid_cinta"]
    #         costo_film_stretch = Articulos.objects.filter(articulo='Film stretch (12 kilos por rollo, 90 rollos por pallet) U$D 2,85 x kilo').values("costo_ars")[0]["costo_ars"]
    #         costo_cinta = Articulos.objects.filter(articulo='Cinta').values("costo_ars")[0]["costo_ars"]
    #         value_mp_cinta_stretch = (Decimal(1)/ins_unid_film*costo_film_stretch+Decimal(1)/ins_unid_cinta*costo_cinta)*cant_unidades
    #         # 3.11 codigo_caja
    #         if request.POST["codigo"] != None:
    #             costo_codigo_caja = Articulos.objects.filter(articulo=request.POST["codigo"]).values("costo_ars")[0]["costo_ars"]
    #             value_codigo_caja = (costo_codigo_caja+1/velocidad*costo_hora_operario)*cant_unidades
    #         # 3.12 costo_embalaje_esp
    #         if request.POST["embalaje"] != None:
    #             costo_embalaje_esp = Articulos.objects.filter(articulo=request.POST["embalaje"]).values("costo_ars")[0]["costo_ars"]
    #             velocidad_aplicacion_u_h_embalaje_esp = Articulos.objects.filter(articulo=request.POST["embalaje"]).values("velocidad_aplicacion_u_h")[0]["velocidad_aplicacion_u_h"]
    #             value_embalaje = (costo_embalaje_esp+1/velocidad_aplicacion_u_h_embalaje_esp*costo_hora_operario)*cant_unidades
    #         # 3.13 logistica
    #         value_logistica = (costo_mensual_logistica/gral_prod_mens_bsas)*cant_unidades
    #         # 3.14 mantenimiento
    #         value_mantenimiento = (costo_mensual_mantenimiento/gral_prod_mens_bsas)*cant_unidades
    #         # 3.15 disenio_pelicula
    #         if request.POST["disenio"] and (cant_colores != None):
    #             disenio_perc = Decimal(request.POST["disenio_perc"])/100
    #             costo_x_pelicula = Articulos.objects.filter(articulo=request.POST["articulo"]).values("costo_ars")[0]["costo_ars"]
    #             dis_cant_bsas = ProduccionParametros.objects.filter(id=1).values()[0]["dis_cant_bsas"]
    #             dis_cant_mza = ProduccionParametros.objects.filter(id=1).values()[0]["dis_cant_mza"]
    #             value_disenio_pelicula = disenio_perc*(costo_mensual_disenio/(dis_cant_bsas+dis_cant_mza)+costo_x_pelicula)*cant_colores
    #         # 3.16 transporte
    #         if request.POST["transporte"]:
    #             value_transporte = Decimal(request.POST["transporte"])
    #         # 3.17 costo_produccion
    #         value_costo_produccion = value_mp_articulo+value_quemado+value_hh_cambio_screen+value_hh_decoracion+ \
    #                                  value_hh_horno+value_mp_screen+value_hh_screen_cop+value_luz_gas+value_mp_pintura+ \
    #                                  value_sampista+value_rot_impresion+value_rot_horno+value_lavado+value_calco_vitrificable+ \
    #                                  value_mp_cinta_stretch+value_codigo_caja+value_embalaje+value_logistica+value_mantenimiento+ \
    #                                  value_setup+value_disenio_pelicula+value_transporte
    #         # 3.18 costo_mp
    #         value_costo_mp = value_mp_articulo+value_mp_cinta_stretch+value_mp_pintura+value_mp_screen
    #         # 3.19 costo_funcionamiento
    #         finan_costo_funcionamiento = FinanzasParametros.objects.filter(id=1).values()[0]["costo_funcionamiento"]
    #         perc_propio = FinanzasParametros.objects.filter(id=1).values()[0]["perc_propio"]/Decimal(100)
    #         perc_cliente = FinanzasParametros.objects.filter(id=1).values()[0]["perc_cliente"]/Decimal(100)
    #         ventas_estimadas_propio = FinanzasParametros.objects.filter(id=1).values()[0]["ventas_estimadas_propio"]
    #         ventas_estimadas_cliente = FinanzasParametros.objects.filter(id=1).values()[0]["ventas_estimadas_cliente"]
    #         ventas_estimadas_propio_art = FinanzasParametros.objects.filter(id=1).values()[0]["compras_estimadas_propio"]
    #         ventas_estimadas_propio_cliente = ventas_estimadas_propio-ventas_estimadas_propio_art
    #         ventas_estimadas_deco_total = ventas_estimadas_propio_cliente+ventas_estimadas_cliente
    #         perc_costo_cliente = (perc_cliente*finan_costo_funcionamiento)/ventas_estimadas_deco_total
    #         perc_costo_propio = (perc_propio*finan_costo_funcionamiento)/ventas_estimadas_propio_art
    #         costos_propios = value_mp_articulo+rot_horno_propia+rot_impresion_propia
    #         costos_cliente = value_costo_produccion-costos_propios
    #         value_costo_funcionamiento_articulo = perc_costo_propio*costos_propios
    #         value_costo_funcionamiento_decorado = perc_costo_cliente*costos_cliente
    #         value_costo_funcionamiento_total = value_costo_funcionamiento_articulo+value_costo_funcionamiento_decorado
    #         # 3.20 costo_total
    #         value_costo_total = value_costo_produccion+value_costo_funcionamiento_total
    #         # 3.21 costo_unitario fabricacion
    #         value_costo_unitario_fab = value_costo_total/cant_unidades
    #         value_costo_unitario_fab_articulo = (costos_propios+value_costo_funcionamiento_articulo)/cant_unidades
    #         value_costo_unitario_fab_decorado = (costos_cliente+value_costo_funcionamiento_decorado)/cant_unidades
    #         # 3.22 costos con ajustes financieros e impositivos
    #         value_costo_unitario = value_costo_unitario_fab
    #         list_cargos = ["iibb_bsas", "iibb_mza", "hig_seg", "com_ventas", "com_otras"]
    #         for cargo in list_cargos:
    #             if request.POST[cargo]:
    #                 value_costo_unitario += value_costo_unitario_fab/Decimal(1-Decimal(request.POST[cargo])/100)-value_costo_unitario_fab
    #         value_costo_unitario_15 = (value_costo_unitario)*(1+costo_deuda/Decimal(2)) # Calculado para 15 días
    #         value_costo_unitario_30 = (value_costo_unitario)*(1+costo_deuda) # Calculado para 30 días
    #         value_costo_unitario_45 = (value_costo_unitario)*(1+costo_deuda*Decimal(1.5)) # Calculado para 45 días
    #         value_costo_unitario_60 = (value_costo_unitario)*(1+costo_deuda*Decimal(2)) # Calculado para 60 días

    #         # 4. Ya realizados los cálculos, si no saltó error guardo los datos en CosteoSerigrafia
    #         new_costeo_serig = form.save(commit=False)
    #         for field_name, field_value in form.cleaned_data.items():
    #             setattr(new_costeo_serig, field_name, field_value)
    #         new_costeo_serig.tipo_2 = tipo_2
    #         new_costeo_serig.save()

    #         # Luego de guardar lo de CosteoSerigrafia, se guardan los datos en la tabla CotizadorSerigrafia
    #         new_cotizador = CotizadorSerigrafia(mp_articulo=value_mp_articulo, quemado=value_quemado, \
    #                                             hh_cambio_screen=value_hh_cambio_screen, hh_decoracion=value_hh_decoracion, \
    #                                             hh_horno=value_hh_horno, mp_screen=value_mp_screen, hh_screen_cop=value_hh_screen_cop, \
    #                                             luz_gas=value_luz_gas, mp_pintura=value_mp_pintura, sampista=value_sampista, \
    #                                             rot_impresion=value_rot_impresion, rot_horno=value_rot_horno, lavado=value_lavado, \
    #                                             calco_vitrificable=value_calco_vitrificable, mp_cinta_stretch=value_mp_cinta_stretch, \
    #                                             codigo_caja=value_codigo_caja, embalaje=value_embalaje, logistica=value_logistica, \
    #                                             mantenimiento=value_mantenimiento, setup=value_setup, disenio_pelicula=value_disenio_pelicula, \
    #                                             transporte=value_transporte, costo_unitario=value_costo_unitario, \
    #                                             costo_unitario_15=value_costo_unitario_15, costo_unitario_30=value_costo_unitario_30, \
    #                                             costo_unitario_45=value_costo_unitario_45, costo_unitario_60=value_costo_unitario_60, \
    #                                             costo_produccion=value_costo_produccion, costo_mp=value_costo_mp,\
    #                                             costo_funcionamiento_articulo = value_costo_funcionamiento_articulo, \
    #                                             costo_funcionamiento_decorado = value_costo_funcionamiento_decorado, \
    #                                             costo_funcionamiento_total=value_costo_funcionamiento_total, costo_total=value_costo_total,\
    #                                             costo_unitario_fab=value_costo_unitario_fab, created_by=request.user, \
    #                                             costo_unitario_fab_articulo=value_costo_unitario_fab_articulo, costo_unitario_fab_decorado=value_costo_unitario_fab_decorado)
    #         new_cotizador.save()
    #         new_costeo_serig.cotizacion = new_cotizador
    #         new_costeo_serig.save()

    #         messages.success(request, "Envío de datos exitoso")
    #     else:
    #         messages.warning(request, f"Envío de datos fallido: {form.errors}")
            

        # return redirect("costeo_serigrafia")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas']).exists())
def costeo_coating(request):
    if request.method == "GET":
        return render(request, "costeo_coating.html")
    if request.method == "POST":
        # Pendiente validación de formularios
        messages.success(request, "Envío de datos exitoso")
        return render(request, "costeo_coating.html")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas', 'finanzas']).exists())
def cotizaciones_hist(request):
    if request.method == "GET":
        id_list = CosteoSerigrafia.objects.order_by('-id')
        articulos_list = Articulos.objects.all()
        cotizaciones_fields = CosteoSerigrafia._meta.get_fields
        clientes_list = Clientes.objects.all()
        #get users from "admin", "ventas" and "finanzas" groups
        users_list = User.objects.filter(groups__name__in=['admin', 'ventas'])

        #get unique values in articulos_list of tipo_2 attribute
        unique_tipo_2 = set(map(lambda x:x.tipo_2, articulos_list))

        return render(request, "cotizaciones_hist.html", {
            "id_list":id_list,
            "articulos_list":articulos_list,
            "unique_tipo_2":unique_tipo_2,
            "cotizaciones_fields":cotizaciones_fields,
            "clientes_list":clientes_list,
            "users_list":users_list,
        })
    elif request.method == "POST":

        form = CotizadorHistoricoFiltersForm(request.POST)

        if form.is_valid():
            id_list = CosteoSerigrafia.objects.order_by('-id')
            articulos_list = Articulos.objects.all()
            cotizaciones_fields = CosteoSerigrafia._meta.get_fields
            unique_tipo_2 = set(map(lambda x: x.tipo_2, articulos_list))
            costeo_list = id_list
            #get all individual clientes from Clientes table
            clientes_list = Clientes.objects.all()
            users_list = User.objects.filter(groups__name__in=['admin', 'ventas', 'finanzas'])

            if request.POST['articulo'] != '':
                costeo_list = costeo_list.filter(articulo=request.POST['articulo'])
            if request.POST['created_by'] != '':
                costeo_list = costeo_list.filter(created_by__username=request.POST['created_by'])
            if request.POST['cliente'] != '':
                costeo_list = costeo_list.filter(cliente__cliente=request.POST['cliente'])
            if request.POST['tipo_cotiz'] != '':
                costeo_list = costeo_list.filter(tipo_cotiz=request.POST['tipo_cotiz'])
            if request.POST['tipo_2'] != '':
                costeo_list = costeo_list.filter(tipo_2=request.POST['tipo_2'])
            if request.POST['_id'] != '':
                costeo_list = costeo_list.filter(id=request.POST['_id'])
            if request.POST['fecha_desde'] != '' and request.POST['fecha_hasta'] != '':
                # Obtener las fechas desde el formulario
                fecha_desde = dt.datetime.strptime(request.POST['fecha_desde'], '%Y-%m-%d').date()
                fecha_hasta = dt.datetime.strptime(request.POST['fecha_hasta'], '%Y-%m-%d').date()

                # Agregar la hora predeterminada
                hora_inicio = time(0, 0, 0)  # 00:00:00
                hora_fin = time(23, 59, 59)  # 23:59:59

                # Combinar las fechas y las horas
                fecha_desde = dt.datetime.combine(fecha_desde, hora_inicio)
                fecha_hasta = dt.datetime.combine(fecha_hasta, hora_fin)
                costeo_list = costeo_list.filter(createdon__gte=fecha_desde, createdon__lte=fecha_hasta)

                #En caso de que solo se ingrese fecha desde, mostrar todas las echas mayor a esa.
            elif request.POST['fecha_desde'] != '':
                fecha_desde = dt.datetime.strptime(request.POST['fecha_desde'], '%Y-%m-%d').date()
                hora_inicio = time(0, 0, 0)
                fecha_desde = dt.datetime.combine(fecha_desde, hora_inicio)
                costeo_list = costeo_list.filter(createdon__gte=fecha_desde)

            #En caso de que solo se ingrese fecha hasta, mostrar todas las echas menor a esa.
            elif request.POST['fecha_hasta'] != '':
                fecha_hasta = dt.datetime.strptime(request.POST['fecha_hasta'], '%Y-%m-%d').date()
                hora_fin = time(23, 59, 59)
                fecha_hasta = dt.datetime.combine(fecha_hasta, hora_fin)
                costeo_list = costeo_list.filter(createdon__lte=fecha_hasta)

            serig_fields = [field for field in CosteoSerigrafia._meta.get_fields() if
                            field.concrete]
            cotiz_fields = [field for field in CotizadorSerigrafia._meta.get_fields() if field.concrete]
            costo_deuda = FinanzasParametros.objects.filter(id=1).values()[0]["costo_deuda"]/100

            context = {
                'id_list': id_list, 
                'articulos_list': articulos_list,
                'unique_tipo_2': unique_tipo_2,
                'cotizaciones_fields': cotizaciones_fields,
                "costeo_list":costeo_list,
                "serig_fields":serig_fields,
                "cotiz_fields":cotiz_fields,
                "costo_deuda":costo_deuda,
                "clientes_list":clientes_list,
                "users_list":users_list,
            } # se lleva el contenido solicitado
            return render(request, "cotizaciones_hist.html", context)       # Obtener los resultados finales
        else:
            print(f"Envío de datos fallido: {form.errors}")
            id_list = CosteoSerigrafia.objects.order_by('-id')
            articulos_list = Articulos.objects.all()
            cotizaciones_fields = CosteoSerigrafia._meta.get_fields
            unique_tipo_2 = set(map(lambda x: x.tipo_2, articulos_list))
            costeo_list = id_list
            form = CotizadorHistoricoFiltersForm(request.POST)
            serig_fields = [field for field in CosteoSerigrafia._meta.get_fields() if
                            field.concrete]
            cotiz_fields = [field for field in CotizadorSerigrafia._meta.get_fields() if field.concrete]
            clientes_list = Clientes.objects.all()
            users_list = User.objects.filter(groups__name__in=['admin', 'ventas', 'finanzas'])

            context = {'form':form,
                       'id_list': id_list, 'articulos_list': articulos_list,
                       'unique_tipo_2': unique_tipo_2, 'cotizaciones_fields': cotizaciones_fields,
                       "costeo_list": costeo_list,
                       "serig_fields": serig_fields,
                       "cotiz_fields": cotiz_fields,
                       "clientes_list": clientes_list,
                       "users_list": users_list,
                       }  # se lleva el contenido solicitado
            return render(request, "cotizaciones_hist.html", context)       # Obtener los resultados finales

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas', 'finanzas']).exists())
def calcular_ganancia(request):
    if request.method == "POST":
        costeo_id = request.POST["costeo_id"]
        porcentaje1 = float(request.POST.get("porcentaje_ganancia", 0))
        porcentaje2 = float(request.POST.get("porcentaje_ganancia2", 0))
        #obtener el objeto de costeo serigrafia con el id
        costeo = CosteoSerigrafia.objects.get(id=costeo_id)
        porcentaje_decorado = 1+(porcentaje1/100)
        porcentaje_articulo = 1+(porcentaje2/100)
        costo_deuda = FinanzasParametros.objects.filter(id=1).values()[0]["costo_deuda"]/100
        porcentaje1 = str(int(porcentaje1)) + "%"
        porcentaje2 = str(int(porcentaje2)) + "%"
        return render(request, "resultado_ganancia.html", {"porcentaje1":porcentaje1, "porcentaje2":porcentaje2, "costo_deuda":costo_deuda, "costeo": costeo, "porcentaje_decorado": porcentaje_decorado, "porcentaje_articulo": porcentaje_articulo})
    return render(request, "cotizaciones_hist.html")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'ventas']).exists())
def nuevo_cliente(request):
    if request.method == 'POST':
        form = NuevoClienteForm(request.POST)

        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.created_by = request.user
            cliente.save()
            return redirect('cotizacion_menu')  # Redirige a la página deseada después de guardar el cliente
        else:
            print(f"Envío de datos fallido: {form.errors}")
            messages.warning(request, 'Error al guardar el cliente: {}'.format(form.errors))
    else:
        form = NuevoClienteForm()  # Crear una instancia vacía del formulario para renderizarlo en la página

    return render(request, 'nuevo_cliente.html', {'form': form})  # Pasar el formulario al contexto del renderizado

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'compras', 'ventas', 'finanzas']).exists())
def compras(request):
    if request.method == "GET":
        return render(request, "compras.html")
    if request.method == "POST":
        if request.FILES.get("file_compras") is not None:
        # Si está subiendo el archivo
            # Primero borro lo que está subido
            Articulos.objects.all().delete()
            # Ya borrado, subo lo nuevo
            file = request.FILES["file_compras"]
            df_main = pd.DataFrame()
            file_xl = load_workbook(file)
            for sheet in file_xl.sheetnames:
                if "aux" not in sheet:
                    df = pd.read_excel(file, sheet)
                    df_main = pd.concat([df_main, df], ignore_index=True)
            df_main = df_main.replace(np.nan, None)
            costo_ars_sinajuste = df_main["costo_ars"]
            costo_deuda = FinanzasParametros.objects.filter(id=1).values()[0]["costo_deuda"]

            try:
                df_main['id'] = df_main.index + 1  # Sumar 1 para evitar índice 0
                for index, (ART, UN, C_A, C_U, C_E, TIPO, TIPO_2, C_P, PESO, DIAM, ASA, CONS_G, CONS_STD, VEL, CSA) \
                in enumerate(zip(df_main.articulo, df_main.unidad, df_main.costo_ars, df_main.costo_usd, \
                    df_main.costo_eur, df_main.tipo, df_main.tipo_2, df_main.condicion_pago_dias, df_main.peso_kg,\
                    df_main.diametro_mm, df_main.asa, df_main.consumo_g_cm2, df_main.consumo_estandar,\
                    df_main.velocidad_aplicacion_u_h, costo_ars_sinajuste), start=1):

                    if C_P != None:
                        C_A = Decimal(C_A/(1+(float(costo_deuda)/30)*(C_P/100)))
                    new_articulos = Articulos(id=index,
                                                articulo=ART, unidad=UN, \
                                            costo_ars=C_A, costo_usd=C_U, \
                                            costo_eur=C_E, tipo=TIPO, \
                                            tipo_2 = TIPO_2, condicion_pago_dias=C_P, \
                                            peso_kg=PESO, diametro_mm=DIAM, \
                                            asa=ASA, consumo_g_cm2=CONS_G, \
                                            consumo_estandar=CONS_STD, velocidad_aplicacion_u_h=VEL, costo_ars_sinajuste=CSA,\
                                            created_by=request.user)
                    new_articulos.save()

                # Pendiente validación de formularios
                messages.success(request, "Envío de datos exitoso")
            except:
                messages.warning(request, "WARNING: Envío de datos fallido. Formato erróneo")

            return render(request, "compras.html")      

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'compras', 'ventas', 'finanzas']).exists())
def compras_detalles(request):
    # Tabla de articulos
    # Vidrio - Tipo: contains 'VIDRIO'
    vidrio_sbtr = ["VIDRIO"]
    vidrio_list = Articulos.objects.filter(reduce(operator.or_, (Q(tipo__contains = i) for i in vidrio_sbtr)))
    articulos_fields = Articulos._meta.get_fields
    # Pintura - Tipo: contains 'TINTA', 'PINTURA'
    pintura_sbtr = ["TINTA", "PINTURA"]
    pintura_list = Articulos.objects.filter(reduce(operator.or_, (Q(tipo__contains = i) for i in pintura_sbtr)))
    # Otros - Tipo: not contains 'VIDRIO', 'TINTA', 'PINTURA'
    otros_sbtr = [*vidrio_sbtr, *pintura_sbtr]
    otros_list = Articulos.objects.exclude(reduce(operator.or_, (Q(tipo__contains = i) for i in otros_sbtr)))

    return render(request, "compras_detail.html", {
        "articulos_fields":articulos_fields,
        "vidrio_list":vidrio_list,
        "pintura_list":pintura_list,
        "otros_list":otros_list
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'rrhh', 'finanzas']).exists())
def rrhh(request):
    if request.method == "GET":
        return render(request, "rrhh.html")
    if request.method == "POST":
        if request.FILES.get("costo_hh") is not None:
            # Si está subiendo el archivo
            # Primero borro lo que está subido
            RecursosHumanos.objects.all().delete()
            # Ya borrado, subo lo nuevo
            file = request.FILES["costo_hh"]
            df_main = pd.DataFrame()
            file_xl = load_workbook(file)
            for sheet in file_xl.sheetnames:
                if "aux" not in sheet:
                    df = pd.read_excel(file, sheet)
                    df_main = pd.concat([df_main, df], ignore_index=True)
            df_main = df_main.replace(np.nan, None)

            # *** SOLUCION 2 ***
            try:
                df_main['id'] = df_main.index + 1  # Sumar 1 para evitar índice 0
                for index, (AP, NOM, PUES, C_H_T, C_M_T) in enumerate(
                        zip(df_main.apellido, df_main.nombre, df_main.puesto,
                            df_main.costo_hora_total, df_main.costo_mensual_total), start=1):
                    new_rrhh = RecursosHumanos(
                        id=index,  # Usar el índice de Pandas como "id" en Django
                        apellido=AP,
                        nombre=NOM,
                        puesto=PUES,
                        costo_hora_total=C_H_T,
                        costo_mensual_total=C_M_T,
                        created_by=request.user
                    )
                    new_rrhh.save()

                # Pendiente validación de formularios
                messages.success(request, "Envío de datos exitoso")
                return render(request, "rrhh.html")
            except:
                messages.warning(request, "WARNING: Envío de datos fallido")
                return render(request, "rrhh.html")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'rrhh', 'finanzas']).exists())
def rrhh_detalles(request):
    rrhh_fields = RecursosHumanos._meta.get_fields
    rrhh_list = RecursosHumanos.objects.all()

    return render(request, "rrhh_detail.html", {
        "rr_hh_fields": rrhh_fields,
        "rr_hh_list": rrhh_list,
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin', 'finanzas', 'ventas']).exists())
def finanzas(request):
    if request.method == "GET":
        #get first instance of finanzas
        finanzas = FinanzasParametros.objects.first()
        #fill the form in finanzas.html with the values of the first instance
        return render(request, "finanzas.html", {
                    "finanzas": finanzas
                })

    if request.method == "POST":
        request.POST._mutable = True
        request.POST["created_by"] = request.user
        request.POST._mutable = False
        form = FinanzasParametrosForm(request.POST)
        if form.is_valid():
            finan_param = FinanzasParametros.objects.get(id=1)
            form_inputs = list(request.POST.keys())
            form_inputs.remove("csrfmiddlewaretoken")
            for input in form_inputs:
                # Hay que saltear la key que trae el post debido al csrf token del formulario
                if input == "costo_deuda":
                    old_value = getattr(finan_param, input)
                    new_value = request.POST[input]
                    if (old_value != new_value) and (request.POST[input]!=""):
                        # Si el valor cambió, hay que actualizar el campo costo_ars de la tabla Articulos
                        # Primero, obtener todos los articulos
                        articulos = Articulos.objects.all()
                        # Luego, actualizar el campo costo_ars de cada articulo
                        for articulo in articulos:
                            if articulo.condicion_pago_dias != None:
                                articulo.costo_ars = articulo.costo_ars_sinajuste/(1 + (Decimal(new_value)/30)*(Decimal(articulo.condicion_pago_dias)/100))
                                articulo.save()

                else:
                    new_value = request.POST[input]
                if new_value != "":
                    # El input del formulario, tiene el mismo nombre que el campo en la base de datos
                    setattr(finan_param, input, new_value)
                    finan_param.save()
            messages.success(request, "Envío de datos exitoso")
        else:
            messages.warning(request, f"Envío de datos fallido: {form.errors}")
        return redirect("finanzas")

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['admin']).exists())
def produccion(request):
    if request.method == "GET":
        produccion = ProduccionParametros.objects.first()
        return render(request, "produccion.html", {
                    "produccion": produccion
                })

    if request.method == "POST":
        request.POST._mutable = True
        request.POST["created_by"] = request.user
        request.POST._mutable = False
        form = ProduccionParametrosForm(request.POST)
        if form.is_valid():
            prod_param = ProduccionParametros.objects.get(id=1)
            form_inputs = list(request.POST.keys())
            form_inputs.remove("csrfmiddlewaretoken")
            for input in form_inputs:
                new_value = request.POST[input]
                if new_value != "":
                    setattr(prod_param, input, new_value)
                    prod_param.save()
            messages.success(request, "Envío de datos exitoso")
        else:
            messages.warning(request, f"Envío de datos fallido: {form.errors}")
    return redirect("produccion")


def descargar_excel(request):
    # Obtener todos los objetos de Articulos
    articulos = Articulos.objects.all()

    # Crear un DataFrame de pandas con los datos de los objetos Articulos
    data = {
        'articulo': [articulo.articulo for articulo in articulos],
        'unidad': [articulo.unidad for articulo in articulos],
        'costo_ars_sinajuste': [articulo.costo_ars_sinajuste for articulo in articulos],
        'costo_ars': [articulo.costo_ars for articulo in articulos],
        'costo_usd': [articulo.costo_usd for articulo in articulos],
        'costo_eur': [articulo.costo_eur for articulo in articulos],
        'tipo': [articulo.tipo for articulo in articulos],
        'tipo_2': [articulo.tipo_2 for articulo in articulos],
        'condicion_pago_dias': [articulo.condicion_pago_dias for articulo in articulos],
        'peso_kg': [articulo.peso_kg for articulo in articulos],
        'diametro_mm': [articulo.diametro_mm for articulo in articulos],
        'asa': [articulo.asa for articulo in articulos],
        'consumo_g_cm2': [articulo.consumo_g_cm2 for articulo in articulos],
        'consumo_estandar': [articulo.consumo_estandar for articulo in articulos],
        'velocidad_aplicacion_u_h': [articulo.velocidad_aplicacion_u_h for articulo in articulos],
    }
    df = pd.DataFrame(data)

    # Generar el archivo Excel en memoria
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Preparar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=articulos.xlsx'

    # Guardar el libro de trabajo de Excel en la respuesta HTTP
    workbook.save(response)

    return response